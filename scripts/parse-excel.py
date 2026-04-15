#!/usr/bin/env python3
"""
parse-excel.py
Reads the "Ace Formula" sheet from the source Excel file and writes
data/historical.json with one entry per trading day for all 5 indices.

Usage:
    python scripts/parse-excel.py [--excel PATH]

Defaults to searching for the xlsx file in the parent directory.
"""

from __future__ import annotations

import json
import math
import os
import sys
import glob
import argparse
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Column layout (0-indexed) from "Ace Formula" sheet
# Row 4 is the header row; data starts at row 5 (index 4 in 0-based).
# ---------------------------------------------------------------------------
INDEX_MAP = {
    "NIFTY_50": {
        "date_col": 1, "close_col": 2, "pb_col": 3, "pe_col": 4,
        "display": "NIFTY 50",
    },
    "NIFTY_BANK": {
        "date_col": 9, "close_col": 10, "pb_col": 11, "pe_col": 12,
        "display": "NIFTY BANK",
    },
    "NIFTY_IT": {
        "date_col": 41, "close_col": 42, "pb_col": 43, "pe_col": 44,
        "display": "NIFTY IT",
    },
    "NIFTY_MIDCAP_150": {
        "date_col": 57, "close_col": 58, "pb_col": 59, "pe_col": 60,
        "display": "Nifty Midcap 150",
    },
    "NIFTY_SMALLCAP_250": {
        "date_col": 65, "close_col": 66, "pb_col": 67, "pe_col": 68,
        "display": "Nifty Smallcap 250",
    },
    "NIFTY_MICROCAP_250": {
        "date_col": 33, "close_col": 34, "pb_col": 35, "pe_col": 36,
        "display": "Nifty Microcap 250",
    },
    "NIFTY_PSU_BANK": {
        "date_col": 49, "close_col": 50, "pb_col": 51, "pe_col": 52,
        "display": "NIFTY PSU BANK",
    },
}

DATA_START_ROW = 5  # 1-indexed row where actual data begins


def parse_date(cell_value) -> str | None:
    """Convert a cell value (datetime, date, or string) to ISO 'YYYY-MM-DD'."""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date().isoformat()
    if hasattr(cell_value, "isoformat"):  # date object
        return cell_value.isoformat()
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(cell_value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def to_float(val) -> float | None:
    """Safely convert a cell value to float, returning None on failure."""
    if val is None:
        return None
    try:
        f = float(val)
        return None if math.isnan(f) or math.isinf(f) else round(f, 4)
    except (TypeError, ValueError):
        return None


def build_metrics(close: float | None, pe: float | None, pb: float | None) -> dict:
    implied_eps = round(close / pe, 2) if (close and pe and pe != 0) else None
    implied_bv = round(close / pb, 2) if (close and pb and pb != 0) else None
    return {
        "close": close,
        "pe": pe,
        "pb": pb,
        "impliedEPS": implied_eps,
        "impliedBV": implied_bv,
    }


def find_excel(base_dir: str) -> str:
    """Auto-locate the xlsx file in base_dir or its parent."""
    for directory in [base_dir, os.path.dirname(base_dir)]:
        matches = glob.glob(os.path.join(directory, "20260128*.xlsx"))
        if matches:
            return matches[0]
    raise FileNotFoundError(
        "Could not find the source Excel file. "
        "Pass --excel PATH to specify it explicitly."
    )


def main():
    parser = argparse.ArgumentParser(description="Parse Excel → historical.json")
    parser.add_argument(
        "--excel",
        default=None,
        help="Path to the source xlsx file (auto-detected if omitted)",
    )
    args = parser.parse_args()

    # Resolve paths relative to the project root (one level up from scripts/)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    output_path = os.path.join(project_root, "data", "historical.json")

    excel_path = args.excel or find_excel(project_root)
    print(f"Reading: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if "Ace Formula" not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        print(f"ERROR: Sheet 'Ace Formula' not found. Available: {available}")
        sys.exit(1)

    ws = wb["Ace Formula"]
    rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
    print(f"Found {len(rows)} data rows starting at row {DATA_START_ROW}")

    result = []
    skipped = 0

    for row in rows:
        # Use NIFTY_50 date column as the canonical date for this row
        raw_date = row[INDEX_MAP["NIFTY_50"]["date_col"]] if len(row) > INDEX_MAP["NIFTY_50"]["date_col"] else None
        date_str = parse_date(raw_date)
        if not date_str:
            skipped += 1
            continue

        entry = {"date": date_str}

        for key, cols in INDEX_MAP.items():
            def safe_get(col_idx):
                return row[col_idx] if len(row) > col_idx else None

            close = to_float(safe_get(cols["close_col"]))
            pb = to_float(safe_get(cols["pb_col"]))
            pe = to_float(safe_get(cols["pe_col"]))
            entry[key] = build_metrics(close, pe, pb)

        result.append(entry)

    # Sort by date ascending, deduplicate
    result.sort(key=lambda r: r["date"])
    seen = set()
    deduped = []
    for r in result:
        if r["date"] not in seen:
            seen.add(r["date"])
            deduped.append(r)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(deduped, f, indent=2)

    print(f"Written {len(deduped)} rows ({skipped} skipped) → {output_path}")
    if deduped:
        print(f"Date range: {deduped[0]['date']} → {deduped[-1]['date']}")
        # Spot-check last row
        last = deduped[-1]
        n50 = last.get("NIFTY_50", {})
        print(f"Last row NIFTY_50: close={n50.get('close')}, pe={n50.get('pe')}, pb={n50.get('pb')}")


if __name__ == "__main__":
    main()
