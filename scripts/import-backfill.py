#!/usr/bin/env python3
"""
import-backfill.py
One-off script to merge Excel backfill data into historical.json.

Reads an Excel file with columns: Date, Close, P/E, P/B (any capitalisation).
Derives impliedEPS = close / pe and impliedBV = close / pb.
Merges data into historical.json by date — updates existing rows and appends
new ones (with null for all other indices). Null-pads existing rows that have
no backfill coverage for the given index key.

Usage (single sheet):
    python scripts/import-backfill.py --file backfill.xlsx --sheet "Nifty Auto" --key NIFTY_AUTO

Usage (multiple keys from multiple sheets in one file):
    python scripts/import-backfill.py --file backfill.xlsx --multi \
      "Nifty Auto:NIFTY_AUTO" \
      "Nifty Financial Services:NIFTY_FIN_SERVICE"
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ALL_KEYS = [
    "NIFTY_50", "NIFTY_BANK", "NIFTY_500", "NIFTY_MIDCAP_150", "NIFTY_SMALLCAP_250",
    "NIFTY_IT", "NIFTY_PSU_BANK", "NIFTY_MICROCAP_250",
    "NIFTY_AUTO", "NIFTY_FIN_SERVICE", "NIFTY_REALTY",
    "NIFTY_METAL", "NIFTY_CAPITAL_MARKETS", "NIFTY_INDIA_DEFENCE",
]

NULL_ENTRY = {"close": None, "pe": None, "pb": None, "impliedEPS": None, "impliedBV": None}


def safe_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(str(val).replace(",", "").strip())
        return None if math.isnan(f) or math.isinf(f) else round(f, 4)
    except (TypeError, ValueError):
        return None


def parse_date(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def find_col(headers: list[str], candidates: list[str]) -> Optional[int]:
    lowers = [h.lower().strip() for h in headers]
    for c in candidates:
        try:
            return lowers.index(c.lower())
        except ValueError:
            continue
    return None


def read_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Skip leading blank rows (some sheets have an empty row before headers)
    first = next((i for i, r in enumerate(rows) if any(c is not None for c in r)), 0)
    rows = rows[first:]

    headers = [str(c) if c is not None else "" for c in rows[0]]
    col_date  = find_col(headers, ["date", "ndp_date"])
    col_close = find_col(headers, ["close", "closing price", "price"])
    col_pe    = find_col(headers, ["p/e", "pe", "p_e", "price/earnings"])
    col_pb    = find_col(headers, ["p/b", "pb", "p_b", "price/book", "price_bv", "price/bv"])

    if col_date is None or col_close is None:
        print(f"ERROR: Could not find Date or Close columns in '{sheet_name}'. Headers: {headers}")
        sys.exit(1)

    records = []
    for row in rows[1:]:
        date_iso = parse_date(row[col_date] if col_date < len(row) else None)
        if not date_iso:
            continue
        close = safe_float(row[col_close] if col_close < len(row) else None)
        pe    = safe_float(row[col_pe]    if col_pe    is not None and col_pe    < len(row) else None)
        pb    = safe_float(row[col_pb]    if col_pb    is not None and col_pb    < len(row) else None)

        implied_eps = round(close / pe, 2) if (close and pe and pe != 0) else None
        implied_bv  = round(close / pb, 2) if (close and pb and pb != 0) else None

        records.append({
            "date": date_iso,
            "close": close,
            "pe": pe,
            "pb": pb,
            "impliedEPS": implied_eps,
            "impliedBV": implied_bv,
        })
    return records


def merge(historical_path: str, key: str, records: list[dict]) -> None:
    with open(historical_path) as f:
        rows: list[dict] = json.load(f)

    row_map: dict[str, dict] = {r["date"]: r for r in rows}
    backfill_map: dict[str, dict] = {r["date"]: r for r in records}

    # Null-pad existing rows that have no backfill for this key
    for date_iso, row in row_map.items():
        if key not in row or row[key] is None:
            if date_iso in backfill_map:
                b = backfill_map[date_iso]
                row[key] = {
                    "close": b["close"], "pe": b["pe"], "pb": b["pb"],
                    "impliedEPS": b["impliedEPS"], "impliedBV": b["impliedBV"],
                }
            else:
                row[key] = dict(NULL_ENTRY)

    # Append new dates from backfill not in historical
    new_dates = set(backfill_map) - set(row_map)
    for date_iso in sorted(new_dates):
        b = backfill_map[date_iso]
        new_row: dict = {"date": date_iso}
        for k in ALL_KEYS:
            if k == key:
                new_row[k] = {
                    "close": b["close"], "pe": b["pe"], "pb": b["pb"],
                    "impliedEPS": b["impliedEPS"], "impliedBV": b["impliedBV"],
                }
            elif k in row_map.get(date_iso, {}):
                new_row[k] = row_map[date_iso][k]
            else:
                new_row[k] = dict(NULL_ENTRY)
        row_map[date_iso] = new_row

    merged = sorted(row_map.values(), key=lambda r: r["date"])

    with open(historical_path, "w") as f:
        json.dump(merged, f, indent=2)

    print(f"[OK] Merged {len(records)} backfill records for {key} → {len(merged)} total rows in {historical_path}")
    filled  = sum(1 for r in records if r["close"] is not None)
    print(f"     {filled}/{len(records)} records have a non-null close price")


def main():
    parser = argparse.ArgumentParser(description="Import Excel backfill data into historical.json")
    parser.add_argument("--file", required=True, help="Path to Excel file (.xlsx)")
    parser.add_argument("--sheet", default=None, help="Sheet name (single-key mode)")
    parser.add_argument("--key",   default=None, help="Internal index key, e.g. NIFTY_AUTO")
    parser.add_argument("--multi", nargs="*", metavar="SHEET:KEY",
                        help="Multiple sheet:key pairs, e.g. 'Nifty Auto:NIFTY_AUTO'")
    args = parser.parse_args()

    script_dir   = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    historical_path = os.path.join(project_root, "data", "historical.json")

    if not os.path.exists(historical_path):
        print(f"ERROR: {historical_path} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)

    pairs: list[tuple[str, str]] = []

    if args.multi:
        for item in args.multi:
            if ":" not in item:
                print(f"ERROR: --multi items must be 'SheetName:INDEX_KEY', got: {item}")
                sys.exit(1)
            sheet, key = item.rsplit(":", 1)
            pairs.append((sheet.strip(), key.strip()))
    elif args.sheet and args.key:
        pairs.append((args.sheet, args.key))
    else:
        print("ERROR: provide either --sheet + --key, or --multi SHEET:KEY pairs")
        sys.exit(1)

    for sheet_name, key in pairs:
        if key not in ALL_KEYS:
            print(f"ERROR: Unknown index key '{key}'. Valid keys: {ALL_KEYS}")
            sys.exit(1)
        print(f"\nProcessing sheet='{sheet_name}' → key={key}")
        records = read_sheet(wb, sheet_name)
        print(f"  Read {len(records)} data rows from sheet")
        merge(historical_path, key, records)


if __name__ == "__main__":
    main()
