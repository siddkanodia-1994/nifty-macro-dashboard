#!/usr/bin/env python3
"""
parse-byey.py
Reads the "BY-EY" sheet from the BY-EY Analysis Excel file and writes
data/byey.json with one entry per trading day.

Usage:
    python scripts/parse-byey.py [--excel PATH]

Defaults to searching for the BY-EY xlsx file in the parent directory.
"""

from __future__ import annotations

import json
import math
import os
import sys
import glob
import argparse
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Column indices (0-based) in the "BY-EY" sheet
COL_DATE       = 1   # B: NDP_Date
COL_CLOSE      = 2   # C: NDP_Close
COL_PE         = 3   # D: NDP_Consolidated PE
COL_EY         = 4   # E: Earnings Yield  (decimal, e.g. 0.0476)
COL_BOND_YIELD = 5   # F: Bond Yield      (decimal, e.g. 0.0690)
COL_SPREAD     = 6   # G: BY - EY         (decimal, e.g. 0.0217)

DATA_START_ROW = 2   # Row 1 = header; data starts at row 2 (1-indexed)


def parse_date(cell_value) -> Optional[str]:
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date().isoformat()
    if hasattr(cell_value, "isoformat"):
        return cell_value.isoformat()
    if isinstance(cell_value, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(cell_value.strip(), fmt).date().isoformat()
            except ValueError:
                continue
    return None


def to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(val)
        return None if math.isnan(f) or math.isinf(f) else round(f, 6)
    except (TypeError, ValueError):
        return None


def find_excel(base_dir: str) -> str:
    for directory in [base_dir, os.path.dirname(base_dir)]:
        for pattern in ["*BY-EY*.xlsx", "*BY EY*.xlsx", "*BYEY*.xlsx"]:
            matches = glob.glob(os.path.join(directory, pattern))
            if matches:
                return matches[0]
    raise FileNotFoundError(
        "Could not find the BY-EY Excel file. Pass --excel PATH to specify it."
    )


def main():
    parser = argparse.ArgumentParser(description="Parse BY-EY Excel → byey.json")
    parser.add_argument("--excel", default=None, help="Path to the BY-EY xlsx file")
    args = parser.parse_args()

    script_dir   = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    output_path  = os.path.join(project_root, "data", "byey.json")

    excel_path = args.excel or find_excel(project_root)
    print(f"Reading: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "BY" in name.upper() and "EY" in name.upper():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
        print(f"WARNING: Could not find BY-EY sheet. Using first sheet: {sheet_name}")
    else:
        print(f"Using sheet: '{sheet_name}'")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
    print(f"Found {len(rows)} data rows starting at row {DATA_START_ROW}")

    result = []
    skipped = 0

    for row in rows:
        if len(row) <= COL_SPREAD:
            skipped += 1
            continue

        date_str = parse_date(row[COL_DATE])
        if not date_str:
            skipped += 1
            continue

        close      = to_float(row[COL_CLOSE])
        pe         = to_float(row[COL_PE])
        ey         = to_float(row[COL_EY])
        bond_yield = to_float(row[COL_BOND_YIELD])
        spread     = to_float(row[COL_SPREAD])

        # Derive EY from PE if not present
        if ey is None and pe and pe != 0:
            ey = round(1 / pe, 6)
        # Derive spread from components if not present
        if spread is None and bond_yield is not None and ey is not None:
            spread = round(bond_yield - ey, 6)

        result.append({
            "date":      date_str,
            "close":     close,
            "pe":        pe,
            "ey":        ey,
            "bondYield": bond_yield,
            "spread":    spread,
        })

    # Sort by date ascending, deduplicate
    result.sort(key=lambda r: r["date"])
    seen: set[str] = set()
    deduped = []
    for r in result:
        if r["date"] not in seen:
            seen.add(r["date"])
            deduped.append(r)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(deduped, f, indent=2)

    print(f"Written {len(deduped)} rows ({skipped} skipped) → {output_path}")
    if deduped:
        print(f"Date range: {deduped[0]['date']} → {deduped[-1]['date']}")
        last = deduped[-1]
        print(f"Last row: pe={last['pe']}, ey={last['ey']}, bondYield={last['bondYield']}, spread={last['spread']}")


if __name__ == "__main__":
    main()
