import openpyxl, json
from datetime import datetime

WB_PATH   = "../New File.xlsm"
JSON_PATH = "data/historical.json"

wb   = openpyxl.load_workbook(WB_PATH, data_only=True)
data = json.load(open(JSON_PATH))
by_date = {r["date"]: r for r in data}

def to_iso(cell):
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    return str(cell)[:10]

def fill_gaps(sheet, key):
    rows = list(sheet.iter_rows(values_only=True))[1:]
    filled = 0
    for row in rows:
        if row[1] is None:
            continue
        date = to_iso(row[1])
        if date not in by_date:
            continue
        entry = by_date[date].setdefault(key, {
            "close": None, "pe": None, "pb": None,
            "impliedEPS": None, "impliedBV": None
        })
        if entry.get("close")      is None and row[2] is not None: entry["close"]     = row[2]; filled += 1
        if entry.get("pe")         is None and row[3] is not None: entry["pe"]        = row[3]
        if entry.get("pb")         is None and row[4] is not None: entry["pb"]        = row[4]
        if entry.get("impliedEPS") is None and row[5] is not None: entry["impliedEPS"] = row[5]
        if entry.get("impliedBV")  is None and row[6] is not None: entry["impliedBV"]  = row[6]
    print(f"  {key}: filled {filled} gap rows")

print("Filling Midcap 150...")
fill_gaps(wb["NIFTY MIDCAP 150_Historical_PE_"], "NIFTY_MIDCAP_150")

print("Filling Smallcap 250...")
fill_gaps(wb["Smallcap 250"], "NIFTY_SMALLCAP_250")

print("Filling Microcap 250...")
fill_gaps(wb["Microcap 250"], "NIFTY_MICROCAP_250")

print("Processing Nifty Private Bank...")
pvt_rows = list(wb["Nifty Private Bank"].iter_rows(values_only=True))[1:]
new_dates = 0
for row in pvt_rows:
    if row[1] is None:
        continue
    date = to_iso(row[1])
    entry = {
        "close":     row[2],
        "pe":        row[3],
        "pb":        row[4],
        "impliedEPS": row[5],
        "impliedBV":  row[6],
    }
    if date in by_date:
        by_date[date]["NIFTY_PVT_BANK"] = entry
    else:
        by_date[date] = {"date": date, "NIFTY_PVT_BANK": entry}
        new_dates += 1
print(f"  NIFTY_PVT_BANK: {new_dates} new date rows added")

sorted_rows = sorted(by_date.values(), key=lambda r: r["date"])
with open(JSON_PATH, "w") as f:
    json.dump(sorted_rows, f, separators=(",", ":"))

print(f"Done. Total rows: {len(sorted_rows)}")

# Spot-checks
row_apr1 = by_date.get("2022-04-01", {})
midcap_pe = row_apr1.get("NIFTY_MIDCAP_150", {}).get("pe")
pvt_close = row_apr1.get("NIFTY_PVT_BANK", {}).get("close")
print(f"Spot-check 2022-04-01 NIFTY_MIDCAP_150.pe = {midcap_pe} (expect ~27.48)")
print(f"Spot-check 2022-04-01 NIFTY_PVT_BANK.close = {pvt_close} (expect ~18831.05)")

new_row = by_date.get("2026-05-22")
print(f"Spot-check 2026-05-22 exists: {new_row is not None}, NIFTY_PVT_BANK: {new_row.get('NIFTY_PVT_BANK', {}).get('close') if new_row else 'N/A'}")
